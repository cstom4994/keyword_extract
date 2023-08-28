# Copyright(c) 2023, KaoruXun(尹喆勋)
# Developed for the "digital intelligence" empowerment rural revitalization
# practice group of North China Electric Power University

import tkinter as tk
from tkinter import filedialog
from tkinter import messagebox
from tkinter import font
from tkinter import ttk
import openpyxl as xl
import jieba
import jieba.analyse
import numpy as np
import tensorflow as tf
from tensorflow.keras.layers import Input, Embedding, GlobalAveragePooling1D, Dense
from tensorflow.keras.models import Model
import re
import pickle
import operator
import matplotlib.pyplot as plt
from wordcloud import WordCloud

# 我知道这很丑
run_good = 0
now_file_path = ""
now_save_count_file_path = ""
now_save_cloud_file_path = ""

# 加载离线jieba分词字典
jieba.set_dictionary("trained_model/dict.txt")


def extract_keywords(text, top_k=9):
    keywords = jieba.analyse.textrank(text, topK=top_k, withWeight=False)
    return keywords


# 使用jieba进行中文分词 并移除标点符号
def tokenize(text):
    # return list(jieba.cut_for_search(text))
    # return [
    #     word for word in jieba.cut(text, cut_all=False) if re.match(r"^[\w]+$", word)
    # ]
    return extract_keywords(text)


def get_corpus(sentences):
    # 对文本进行分词
    tokenized_sentences = [tokenize(sentence) for sentence in sentences]
    return tokenized_sentences


def gen_word2index(words):
    # 建立词汇表
    vocab = sorted(set(words))
    word_to_index = {word: index for index, word in enumerate(vocab)}
    index_to_word = {index: word for word, index in word_to_index.items()}
    vocab_size = len(vocab)
    return vocab, word_to_index


def training(sentence):
    tokenized_sentences = get_corpus(sentence)

    # 将分词后的文本转换为单词列表
    words = []
    for sentence in tokenized_sentences:
        words.extend(sentence)

    # 建立词汇表
    vocab, word_to_index = gen_word2index(words)
    vocab_size = len(vocab)

    # 构建Word2Vec词向量模型
    embedding_dim = 100
    input_target = Input(shape=(1,))
    input_context = Input(shape=(1,))
    embedding_layer = Embedding(vocab_size, embedding_dim)
    target_embed = embedding_layer(input_target)
    context_embed = embedding_layer(input_context)

    # 使用点乘计算两个词向量的相似度
    dot_product = tf.keras.layers.Dot(axes=2)([target_embed, context_embed])

    # 定义神经网络模型
    x = GlobalAveragePooling1D()(dot_product)
    x = Dense(64, activation="relu")(x)
    output = Dense(1, activation="sigmoid")(x)
    model = Model(inputs=[input_target, input_context], outputs=output)

    # 编译模型
    model.compile(
        optimizer="adam", loss="binary_crossentropy", metrics=["accuracy", "mse"]
    )

    # 创建训练数据
    positive_pairs = []
    negative_pairs = []
    for i, sentence in enumerate(tokenized_sentences):
        for j in range(len(sentence)):
            for k in range(j + 1, len(sentence)):
                positive_pairs.append(
                    (word_to_index[sentence[j]], word_to_index[sentence[k]])
                )
                # 随机选择一个负样本
                negative_word_index = np.random.randint(0, vocab_size)
                while negative_word_index == word_to_index[sentence[j]]:
                    negative_word_index = np.random.randint(0, vocab_size)
                negative_pairs.append((word_to_index[sentence[j]], negative_word_index))

    target_words = np.array(
        [pair[0] for pair in positive_pairs + negative_pairs], dtype=np.int32
    )
    context_words = np.array(
        [pair[1] for pair in positive_pairs + negative_pairs], dtype=np.int32
    )
    labels = np.array(
        [1] * len(positive_pairs) + [0] * len(negative_pairs), dtype=np.int32
    )

    # 训练模型
    model.fit(x=[target_words, context_words], y=labels, epochs=10, batch_size=16)

    # 获取单词的词向量
    # word_vectors = embedding_layer.get_weights()[0]

    # 打印词向量结果
    # for word, index in word_to_index.items():
    #     print(f"{word}: {word_vectors[index]}")

    # 保存模型到文件
    model.save("trained_model/rnn.keras")

    # # 保存词向量数据
    # with open("trained_model/word_vectors.pkl", "wb") as f:
    #     pickle.dump(word_vectors, f)

    # # 保存词汇表
    # with open("trained_model/vocab.pkl", "wb") as f:
    #     pickle.dump(vocab, f)

    # 保存词汇索引
    with open("trained_model/word_to_index.pkl", "wb") as f:
        pickle.dump(word_to_index, f)

    print(f"模型词汇索引量为: {len(word_to_index)}")

    return


def test(input_sentence):
    # 加载词汇索引
    with open("trained_model/word_to_index.pkl", "rb") as f:
        word_to_index = pickle.load(f)

    # 加载已保存的模型
    loaded_model = tf.keras.models.load_model("trained_model/rnn.keras")

    loaded_model.summary()

    # 获取 embedding_layer
    embedding_layer = loaded_model.get_layer("embedding")

    # 获取词汇表和词向量
    word_vectors = embedding_layer.get_weights()[0]
    # vocab = list(embedding_layer.get_vocabulary())

    # 使用训练好的模型获取单词的词向量
    # word = "喜欢"
    # if word in word_to_index:
    #     word_index = word_to_index[word]
    #     word_vector = loaded_model.layers[2].get_weights()[0][word_index]
    #     print(f"{word} 的词向量：{word_vector}")
    # else:
    #     print(f"{word} 不在词汇表中。")

    # 提取句子关键词
    def extract_keywords(sentence, word_vectors):
        words = tokenize(sentence)
        word_vectors = [
            word_vectors[word_to_index[word]] for word in words if word in word_to_index
        ]
        if word_vectors:
            sentence_vector = np.mean(word_vectors, axis=0)
            similarities = np.dot(word_vectors, sentence_vector) / (
                np.linalg.norm(word_vectors, axis=1) * np.linalg.norm(sentence_vector)
            )
            top_indices = np.argsort(-similarities)[:8]
            keywords = [words[i] for i in top_indices]
            return keywords
        else:
            return []

    all_keywords = []

    # 测试句子关键词提取

    max_len = len(input_sentence)
    i = 1
    for sentence in input_sentence:
        i = i + 1
        keywords = extract_keywords(sentence, word_vectors)
        # print(f"{sentence} \n关键词: {keywords}")
        all_keywords += keywords
        progress_bar["value"] = (i / max_len) * 35 + 60
        root.update_idletasks()  # 更新窗口

    return all_keywords


def count_word_frequency(text):
    # 使用空格分隔文本中的单词
    words = text
    # 创建一个空字典来存储单词和它们的频率
    word_freq = {}
    # 遍历文本中的每个单词
    for word in words:
        # 去除标点符号
        word = word.strip(".,!?\"'()[]{}:;")
        # 将单词转换为小写，以便忽略大小写
        word = word.lower()
        # 更新字典中的单词频率
        if word in word_freq:
            word_freq[word] += 1
        else:
            word_freq[word] = 1
    return word_freq


def run_keyword_extract():
    if run_good < 3:
        messagebox.showerror("错误", f"无法打开表格 {now_file_path}\n\n请检查文件是否完整或程序是否有读取权限")
        return

    training_sentences = []

    wb = xl.load_workbook(now_file_path)
    # 选择默认的工作表
    sheet = wb.active
    max_row = sheet.max_row
    i = 0
    for row in sheet.iter_rows(
        min_row=2, max_row=max_row, min_col=1, max_col=3, values_only=True
    ):
        i += 1
        comment, comment_time, comment_star = row
        comment = re.sub(r"\n", "", comment)
        training_sentences.append(comment)
        progress_bar["value"] = (i / sheet.max_row) * 60
        root.update_idletasks()  # 更新窗口

    # print(training_sentences)

    # training(training_sentences)

    text = test(training_sentences)

    word_frequency = count_word_frequency(text)

    # 按词频从高到低排序字典
    sorted_word_freq = dict(
        sorted(word_frequency.items(), key=operator.itemgetter(1), reverse=True)
    )

    data = sorted_word_freq

    # 创建一个工作簿和工作表
    workbook = xl.Workbook()
    sheet = workbook.active

    # 将key作为列头写入第一行
    header_row = 1
    column = 1
    for key in data.keys():
        sheet.cell(column=header_row, row=column, value=key)
        column += 1

    # 将value写入相应的单元格
    data_row = 2
    for key, value in data.items():
        column = list(data.keys()).index(key) + 1
        sheet.cell(column=data_row, row=column, value=value)

    # 保存Excel文件
    workbook.save(now_save_count_file_path)
    print(f"Data saved to {now_save_count_file_path}")

    second_pass = ""

    # 打印词语及其出现频率
    for word, freq in sorted_word_freq.items():
        second_pass += word + "\n"

    keywords = extract_keywords(second_pass)
    print("总体关键向量：", keywords)

    wordcloud = WordCloud(
        width=1800,
        height=1800,
        font_path="trained_model/sarasa-mono-sc-regular.ttf",
        background_color="white",
        max_words=500,
        min_font_size=8,
        colormap="viridis",
    ).generate(second_pass)

    # 绘制词云图
    plt.figure(figsize=(8, 8), facecolor=None)
    plt.imshow(wordcloud, interpolation="bilinear")
    plt.axis("off")
    plt.tight_layout(pad=0)

    # 保存词云图为图片文件
    wordcloud.to_file(now_save_cloud_file_path)

    progress_bar["value"] = 95
    root.update_idletasks()  # 更新窗口

    # 显示词云图
    plt.show()

    progress_bar["value"] = 100
    root.update_idletasks()  # 更新窗口

    messagebox.showinfo("成功", "成功完成评论关键词提取")

    return


def browse_file():
    global now_file_path, run_good
    file_path = filedialog.askopenfilename(filetypes=[("Excel Files", "*.xlsx")])
    if file_path:
        try:
            wb = xl.load_workbook(file_path)
            sheet = wb.active
            # 获取最大行数
            max_row = sheet.max_row
            run_good = 1
            # run_button.config(state="normal")
            save_count_button.config(state="normal")
            path_var.set(f"当前关键词提取目标文件: {file_path}")
            now_file_path = file_path
            messagebox.showinfo("成功", f"成功打开表格 {file_path}\n\n已读取最大行数为 {max_row}")
        except:
            run_good = 0
            run_button.config(state="disabled")
            now_file_path = ""
            messagebox.showerror("错误", f"无法打开表格 {file_path}\n\n请检查文件是否完整或程序是否有读取权限")
            pass


def save_count_file():
    global now_save_count_file_path, run_good
    if run_good < 1:
        messagebox.showerror("错误", f"无法打开表格 {now_file_path}\n\n请检查文件是否完整或程序是否有读取权限")
        return
    file_path = filedialog.asksaveasfilename(
        defaultextension=".xlsx", filetypes=[("Excel Files", "*.xlsx")]
    )
    if file_path:
        run_good = 2
        save_cloud_button.config(state="normal")
        now_save_count_file_path = file_path
        count_file_path_var.set(f"关键词统计信息保存路径 f{now_save_count_file_path}")
        print(f"File saved to: {file_path}")
    return


def save_cloud_file():
    global now_save_cloud_file_path, run_good
    if run_good < 2:
        messagebox.showerror("错误", f"无法打开表格 {now_file_path}\n\n请检查文件是否完整或程序是否有读取权限")
        return
    file_path = filedialog.asksaveasfilename(
        defaultextension=".png", filetypes=[("png 图像文件", "*.png")]
    )
    if file_path:
        run_good = 3
        run_button.config(state="normal")
        now_save_cloud_file_path = file_path
        cloud_file_path_var.set(f"关键词云图保存路径 {now_save_cloud_file_path}")
        print(f"File saved to: {file_path}")
    return


def increase_font_size():
    global font_size
    font_size += 2
    label_font.configure(size=font_size)


# 创建主窗口
root = tk.Tk()
root.title("评论关键词提取系统")

font_size = 14
label_font = font.Font(size=font_size)

# 设置窗口大小
window_width = 500
window_height = 400
screen_width = root.winfo_screenwidth()
screen_height = root.winfo_screenheight()
x = (screen_width - window_width) // 2
y = (screen_height - window_height) // 2
root.geometry(f"{window_width}x{window_height}+{x}+{y}")

# 禁止窗口在水平和垂直方向上调整大小
root.resizable(False, False)


# 创建显示框的标签
label = tk.Label(root, text='华北电力大学 "数智化"赋能乡村振兴实践团 社会实践成果', font=label_font)
label.pack(pady=10)

# 创建用于显示路径的标签
path_var = tk.StringVar()
path_var.set("未设置 当前关键词提取目标文件")
path_label = tk.Label(root, textvariable=path_var, wraplength=500)
path_label.pack()

count_file_path_var = tk.StringVar()
count_file_path_var.set("未设置 关键词统计信息保存路径")
count_file_path_label = tk.Label(root, textvariable=count_file_path_var, wraplength=500)
count_file_path_label.pack()

cloud_file_path_var = tk.StringVar()
cloud_file_path_var.set("未设置 关键词云图保存路径")
cloud_file_path_label = tk.Label(root, textvariable=cloud_file_path_var, wraplength=500)
cloud_file_path_label.pack()

# 创建浏览按钮
browse_button = tk.Button(root, text="选择评论信息文件", command=browse_file)
browse_button.pack(pady=5)

save_count_button = tk.Button(
    root, text="选择关键词统计信息保存路径", command=save_count_file, state="disabled"
)
save_count_button.pack(pady=5)

save_cloud_button = tk.Button(
    root, text="选择关键词云图保存路径", command=save_cloud_file, state="disabled"
)
save_cloud_button.pack(pady=5)

run_button = tk.Button(
    root, text="开始进行评论关键词提取", command=run_keyword_extract, state="disabled"
)
run_button.pack(side="bottom", pady=10)

# 创建进度条
progress_bar = ttk.Progressbar(
    root, orient="horizontal", length=400, mode="determinate"
)
progress_bar.pack(side="bottom", pady=10)

# 运行主循环
root.mainloop()
